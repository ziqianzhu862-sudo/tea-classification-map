import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

SOURCE_DIR = Path(r"D:\xwechat_files\wxid_0em68shb88or12_3e5f\msg\file\2026-06")
SOURCE_FILE = next(p for p in SOURCE_DIR.glob("*.xlsx") if "茶叶" in p.name and "产地" in p.name)
OUT_DIR = Path(r"D:\Codex\tea-classification-map\data")

PROVINCE_COORDS = {
    "浙江省": [120.2, 29.2],
    "江苏省": [119.8, 32.9],
    "安徽省": [117.2, 31.8],
    "福建省": [118.0, 26.1],
    "江西省": [116.0, 27.6],
    "湖南省": [112.9, 27.8],
    "湖北省": [112.3, 30.9],
    "河南省": [113.7, 33.9],
    "山东省": [118.4, 36.4],
    "四川省": [103.8, 30.5],
    "重庆市": [107.7, 30.0],
    "贵州省": [106.7, 26.8],
    "云南省": [101.8, 24.9],
    "广东省": [113.2, 23.4],
    "广西壮族自治区": [108.3, 23.8],
    "台湾省": [121.0, 23.8],
    "海南省": [110.2, 19.2],
    "陕西省": [108.9, 34.2],
    "甘肃省": [103.8, 36.1],
}

TEA_COLORS = {
    "绿茶": "#4FB07A",
    "红茶": "#EC7045",
    "青茶": "#54A9A2",
    "白茶": "#CFE0CA",
    "黄茶": "#E6C45B",
    "黑茶": "#D978A8",
}

AREA_COLORS = {
    "江南茶区": "#6FC9B6",
    "华南茶区": "#F28C52",
    "西南茶区": "#A4C7E8",
    "江北茶区": "#D9C95B",
}

wb = load_workbook(SOURCE_FILE, data_only=True, read_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
headers = [str(v).strip() for v in rows[0]]

records = []
for idx, row in enumerate(rows[1:], start=1):
    item = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}
    if not item.get("品种级茶名"):
        continue
    item = {k: ("" if v is None else str(v).strip()) for k, v in item.items()}
    item["id"] = idx
    item["color"] = TEA_COLORS.get(item["茶类"], "#78909C")
    item["areaColor"] = AREA_COLORS.get(item["四大茶区"], "#9CA3AF")
    item["coords"] = PROVINCE_COORDS.get(item["省级"], [105, 32])
    records.append(item)

def count_by(field):
    return [{"name": k, "value": v} for k, v in Counter(r[field] for r in records).most_common()]

province_payload = []
province_groups = defaultdict(list)
for r in records:
    province_groups[r["省级"]].append(r)
for province, items in province_groups.items():
    province_payload.append({
        "name": province,
        "value": len(items),
        "coords": PROVINCE_COORDS.get(province, [105, 32]),
        "teaTypes": count_by_field(items, "茶类") if False else None,
    })

def grouped_counts(items, field):
    return [{"name": k, "value": v} for k, v in Counter(r[field] for r in items).most_common()]

for p in province_payload:
    items = province_groups[p["name"]]
    p["teaTypes"] = grouped_counts(items, "茶类")
    p["areas"] = grouped_counts(items, "四大茶区")

subclasses = []
for (tea_type, subclass), items in defaultdict(list).items():
    pass

subclass_groups = defaultdict(list)
for r in records:
    subclass_groups[(r["茶类"], r["子类"])].append(r)
for (tea_type, subclass), items in subclass_groups.items():
    subclasses.append({
        "teaType": tea_type,
        "name": subclass,
        "count": len(items),
        "color": TEA_COLORS.get(tea_type, "#78909C"),
        "examples": [x["品种级茶名"] for x in items[:5]],
    })

payload = {
    "title": "茗香万象：中国茶叶分类与产地地图",
    "sourceFile": str(SOURCE_FILE),
    "records": records,
    "summary": {
        "total": len(records),
        "teaTypes": count_by("茶类"),
        "areas": count_by("四大茶区"),
        "provinces": count_by("省级"),
        "cities": count_by("市级"),
        "subclasses": sorted(subclasses, key=lambda x: (-x["count"], x["teaType"], x["name"])),
    },
    "colors": {
        "teaTypes": TEA_COLORS,
        "areas": AREA_COLORS,
    },
    "provinces": sorted(province_payload, key=lambda x: -x["value"]),
}

OUT_DIR.mkdir(parents=True, exist_ok=True)
(OUT_DIR / "tea-data.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {len(records)} records from {SOURCE_FILE}")
